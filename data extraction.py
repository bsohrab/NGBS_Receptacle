import openpyxl
import csv
import os
import easygui
import glob
import xlsxwriter
import re
import logging, sys

def retrieveoverview(worksheetOV):
    final_loc = 0
    topic_loc = 0
    climate_zone = 0
    radon_level = 0
    for row_num in range(1,worksheetOV.max_row):
        for col_num in range(1,worksheetOV.max_column):
            # print(worksheetOV.cell(row_num, col_num).value)

            if worksheetOV.cell(row_num, col_num).value == "Final":
                final_loc= [col_num,row_num]
                # print(final_loc)
            elif worksheetOV.cell(row_num, col_num).value == "Overview (Verifier phase)":
                topic_loc = [col_num,row_num]
                # print(topic_loc)
            elif worksheetOV.cell(row_num, col_num).value == "Climate Zone:":
                climate_zone = worksheetOV.cell(row_num+1, col_num).value
                # print(climate_zone)
            elif worksheetOV.cell(row_num, col_num).value == "Radon":
                radon_level = worksheetOV.cell(row_num + 1, col_num).value
                # print(radon_level)
            elif final_loc != 0 and topic_loc != 0 and climate_zone != 0 and radon_level != 0:
                # print("Completed")
                return final_loc, topic_loc, climate_zone, radon_level

def dissect(file):
    try:
        workbook=openpyxl.load_workbook((file), data_only=True)
        worksheet = workbook['Verification Report']
    except:
        return



    header = []
    value = 14
    row = 0
    worksheetYYYY = workbook['Info & Intro']
    if "2015" in worksheetYYYY.cell(1,4).value:
        header.append(tuple(("Year","2015")))
    elif "2020" in worksheetYYYY.cell(1,4).value:
        header.append(tuple(("Year","2020")))



    worksheetOV = workbook['Overview (Verification)']
    final_loc, topic_loc, climate_zone, radon_level = retrieveoverview(worksheetOV)
    finalrowOV = 69
    rowOV = final_loc[1]
    while rowOV < finalrowOV:
        if worksheetOV.cell(rowOV, topic_loc[0]).value is not None:
            header.append(tuple((worksheetOV.cell(rowOV,topic_loc[0]).value, worksheetOV.cell(rowOV,final_loc[0]).value)))
        rowOV += 1


    while value <= 2500:#cellrow < 2500:
        criteria = ""
        cell = "$W$" + str(value)

        RorFcheck = [worksheet["$P" + cell[2:]].value, worksheet["$Q" + cell[2:]].value]
        if RorFcheck[0] is None or "X" in str(RorFcheck[0]) or RorFcheck[1] is None or "X" in str(RorFcheck[1]):
            RorFcheck = None
        elif RorFcheck[0] == True and RorFcheck[1] == False:
            RorFcheck = "Rough"
        elif RorFcheck[0] == False and RorFcheck[1] == True:
            RorFcheck = "Final"
        elif RorFcheck[0] == True and RorFcheck[1] == True:
            RorFcheck = "Rough/Final"
        elif RorFcheck[0] == False and RorFcheck[1] == False:
            RorFcheck = "Error Error Error"

        pointsawarded = worksheet["$O" + cell[2:]].value
        booster = 1
        while pointsawarded is None:
            pointsawarded = worksheet["$O" + str(int(cell[3:]) - booster)].value
            booster += 1
        if isinstance(pointsawarded, int) == False:
            pointsawarded = 1
        elif pointsawarded >0:
            pointsawarded=1
        elif pointsawarded ==0:
            pointsawarded = 0



        errorcheck = [worksheet["$S" + cell[2:]].value, worksheet["$T" + cell[2:]].value]
        if errorcheck[0] == True and errorcheck[1] == False:
            errorcheck = "Points Available - no error"
        elif errorcheck[0] == False and errorcheck[1] == True:
            errorcheck = "Points Unavailable - Error"
        elif errorcheck[0] == True and errorcheck[1] == True:
            errorcheck = "Points Available - Error"
        elif errorcheck[0] == False and errorcheck[1] == False:
            errorcheck = "Points Unavailable - no error"



        try:
            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + worksheet[ \
                "$D" + cell[2:]].value + " " + worksheet["$E" + cell[2:]].value + " " + worksheet[ \
                            "$F" + cell[2:]].value
        except TypeError:
            try:#4
                practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                            worksheet["$D" + cell[2:]].value + " " + worksheet["$E" + cell[2:]].value
            except TypeError:
                try:#3
                    practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                                worksheet["$D" + cell[2:]].value
                except TypeError:
                    try:#3
                        practices = str(worksheet["$B" + cell[2:]].value) + " " +worksheet["$E" + cell[2:]].value + " " + \
                                    worksheet["$F" + cell[2:]].value
                    except TypeError:
                        try:#3
                            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[ \
                                "$D" + cell[2:]].value + " " + \
                                        worksheet["$E" + cell[2:]].value
                        except TypeError:
                            try:
                                practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$E" + cell[2:]].value
                            except TypeError:
                                try:
                                    practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[ \
                                        "$D" + cell[2:]].value + " " + \
                                               worksheet["$E" + cell[2:]].value
                                except TypeError:
                                    try:
                                        practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value
                                    except TypeError:
                                        try:
                                            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$F" + cell[2:]].value
                                        except TypeError:
                                            try:
                                                practices = str(worksheet["$B" + cell[2:]].value)
                                            except:
                                                practices = "No Practice Given"

        #logging.info("Practice: " + practices)

        if RorFcheck is None :
            value += 1
            continue
        elif RorFcheck is not None:
            header.append(tuple((practices, pointsawarded)))
            row += 1
        value += 1




    workbook = openpyxl.load_workbook(('C:\\Users\\spasikhani\\Documents\\mass production test\\45.xlsx'), data_only=True)
    worksheet = workbook['Sheet1']
    inputrow = worksheet.max_row+1
    for topics, headerset in enumerate(header):
        for col_num in range(1,worksheet.max_column):
            if worksheet.cell(1, col_num).value == headerset[0]:
                worksheet.cell(inputrow, col_num).value = headerset[1]
    workbook.save('C:\\Users\\spasikhani\\Documents\\mass production test\\45.xlsx')





#companyname = easygui.enterbox('What is the name of the company', "company name (case insensitive")
#datadirectory = "G:/Division/Lab/Green/Certification/Report Reviews/Standard/" + alpharange(companyname)

# datadirectory = "C:/Users/spasikhani/Desktop/backup copyies to avoid testing on/H042023RR.xlsx"
# datadirectory = "G:/Division/Lab/Green/Certification/Report Reviews/Standard/S-U/Urban NW Homes (WA)/750 South 88th Avenue/H041794FR_REVIEW.xlsx"
logging.basicConfig(stream=sys.stderr, level=logging.DEBUG)
# dissect(datadirectory)

datadirectory = "G:/Division/Lab/Green/Certification/Report Reviews/Standard"
#
# logging.debug(datadirectory)
#
# # for folders in glob.glob(datadirectory+"*/*/*", recursive=True):
# #       print(folders)
#
for file in glob.iglob(my_path, recursive=True):
for  files in glob.glob(datadirectory+"*/*/*FR_REVIEW.xlsx", recursive=True):
    logging.debug(files)
    dissect(files)
for  files in glob.glob(datadirectory+"*/*/*/*FR_REVIEW.xlsx", recursive=True):
    logging.debug(files)
    dissect(files)
for  files in glob.glob(datadirectory+"*/*/*/*/*FR_REVIEW.xlsx", recursive=True):
    logging.debug(files)
    dissect(files)