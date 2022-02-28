import os
import win32com.client as win32
import tkinter as tk
import glob
import openpyxl
import csv
import logging, sys
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule


def mandatoryextraction(value, row, worksheet):
    header = []
    header_csv = []
    while value <= 2500:  # cellrow < 2500:
        criteria = ""
        cell = "$W$" + str(value)

        RorFcheck = [worksheet["$P" + cell[2:]].value, worksheet["$Q" + cell[2:]].value]
        if RorFcheck[0] is None or "X" in str(RorFcheck[0]) or RorFcheck[1] is None or "X" in str(RorFcheck[1]):
            RorFcheck = None
        elif RorFcheck[0] == True and RorFcheck[1] == False:
            RorFcheck = "Rough"
        elif RorFcheck[0] == False and RorFcheck[1] == True:
            RorFcheck = "Final"
        elif RorFcheck[0] == True and RorFcheck[1] == True:
            RorFcheck = "Rough/Final"
        elif RorFcheck[0] == False and RorFcheck[1] == False:
            RorFcheck = "Error Error Error"

        ##

        mandatoryitem = worksheet["$R" + cell[2:]].value
        practicetext = worksheet["$L" + cell[2:]].value
        pointschosen = worksheet["$W" + cell[2:]].value
        # booster = 1
        # while pointsawarded is None:
        #     pointsawarded = worksheet["$O" + str(int(cell[3:]) - booster)].value
        #     original_practicetext = worksheet["$L" + cell[2:]].value
        #     booster += 1

        errorcheck = [worksheet["$S" + cell[2:]].value, worksheet["$T" + cell[2:]].value]
        if errorcheck[0] == True and errorcheck[1] == False:
            errorcheck = "Points Available - no error"
        elif errorcheck[0] == False and errorcheck[1] == True:
            errorcheck = "Points Unavailable - Error"
        elif errorcheck[0] == True and errorcheck[1] == True:
            errorcheck = "Points Available - Error"
        elif errorcheck[0] == False and errorcheck[1] == False:
            errorcheck = "Points Unavailable - no error"

        try:
            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                        worksheet[ \
                            "$D" + cell[2:]].value + " " + worksheet["$E" + cell[2:]].value + " " + worksheet[ \
                            "$F" + cell[2:]].value
        except TypeError:
            try:  # 4
                practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                            worksheet["$D" + cell[2:]].value + " " + worksheet["$E" + cell[2:]].value
            except TypeError:
                try:  # 3
                    practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                                worksheet["$D" + cell[2:]].value
                except TypeError:
                    try:  # 3
                        practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                            "$E" + cell[2:]].value + " " + \
                                    worksheet["$F" + cell[2:]].value
                    except TypeError:
                        try:  # 3
                            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[ \
                                "$D" + cell[2:]].value + " " + \
                                        worksheet["$E" + cell[2:]].value
                        except TypeError:
                            try:
                                practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                                    "$E" + cell[2:]].value
                            except TypeError:
                                try:
                                    practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[ \
                                        "$D" + cell[2:]].value + " " + \
                                                worksheet["$E" + cell[2:]].value
                                except TypeError:
                                    try:
                                        practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                                            "$C" + cell[2:]].value
                                    except TypeError:
                                        try:
                                            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                                                "$F" + cell[2:]].value
                                        except TypeError:
                                            practices = str(worksheet["$B" + cell[2:]].value)
        notes = worksheet["$X" + cell[2:]].value
        if notes is None:
            notes = "None"

        if RorFcheck is None:
            value += 1
            continue
        elif RorFcheck is not None and mandatoryitem is True:
            header.append((practices, mandatoryitem, pointschosen, RorFcheck, errorcheck, practicetext,notes))
            header_csv.append(('mandaory',practices, mandatoryitem, pointschosen, RorFcheck, errorcheck, practicetext, notes))
            row += 1
        value += 1
    return header, header_csv
        # logging.info("Practice: " + practices)

def practiceextraction(value, row, worksheet):
    header = []
    header_csv = []
    while value <= 2500:  # cellrow < 2500:
        criteria = ""
        cell = "$W$" + str(value)

        RorFcheck = [worksheet["$P" + cell[2:]].value, worksheet["$Q" + cell[2:]].value]
        if RorFcheck[0] is None or "X" in str(RorFcheck[0]) or RorFcheck[1] is None or "X" in str(RorFcheck[1]):
            RorFcheck = None
        elif RorFcheck[0] == True and RorFcheck[1] == False:
            RorFcheck = "Rough"
        elif RorFcheck[0] == False and RorFcheck[1] == True:
            RorFcheck = "Final"
        elif RorFcheck[0] == True and RorFcheck[1] == True:
            RorFcheck = "Rough/Final"
        elif RorFcheck[0] == False and RorFcheck[1] == False:
            RorFcheck = "Error Error Error"


        pointsawarded = worksheet["$O" + cell[2:]].value
        practicetext = worksheet["$L" + cell[2:]].value
        pointschosen = worksheet["$W" + cell[2:]].value
        original_practicetext = worksheet["$L" + cell[2:]].value
        booster = 1
        while pointsawarded is None:
            pointsawarded = worksheet["$O" + str(int(cell[3:]) - booster)].value
            original_practicetext = worksheet["$L" + cell[2:]].value
            booster += 1
        # if isinstance(pointsawarded, int) == False:
        #     pointsawarded = 1
        # elif pointsawarded >0:
        #     pointsawarded=1
        # elif pointsawarded ==0:
        #     pointsawarded = 0


        errorcheck = [worksheet["$S" + cell[2:]].value, worksheet["$T" + cell[2:]].value]
        if errorcheck[0] == True and errorcheck[1] == False:
            errorcheck = "Points Available - no error"
        elif errorcheck[0] == False and errorcheck[1] == True:
            errorcheck = "Points Unavailable - Error"
        elif errorcheck[0] == True and errorcheck[1] == True:
            errorcheck = "Points Available - Error"
        elif errorcheck[0] == False and errorcheck[1] == False:
            errorcheck = "Points Unavailable - no error"

        try:
            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                        worksheet[ \
                            "$D" + cell[2:]].value + " " + worksheet["$E" + cell[2:]].value + " " + worksheet[ \
                            "$F" + cell[2:]].value
        except TypeError:
            try:  # 4
                practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                            worksheet["$D" + cell[2:]].value + " " + worksheet["$E" + cell[2:]].value
            except TypeError:
                try:  # 3
                    practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet["$C" + cell[2:]].value + " " + \
                                worksheet["$D" + cell[2:]].value
                except TypeError:
                    try:  # 3
                        practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                            "$E" + cell[2:]].value + " " + \
                                    worksheet["$F" + cell[2:]].value
                    except TypeError:
                        try:  # 3
                            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[ \
                                "$D" + cell[2:]].value + " " + \
                                        worksheet["$E" + cell[2:]].value
                        except TypeError:
                            try:
                                practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                                    "$E" + cell[2:]].value
                            except TypeError:
                                try:
                                    practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[ \
                                        "$D" + cell[2:]].value + " " + \
                                                worksheet["$E" + cell[2:]].value
                                except TypeError:
                                    try:
                                        practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                                            "$C" + cell[2:]].value
                                    except TypeError:
                                        try:
                                            practices = str(worksheet["$B" + cell[2:]].value) + " " + worksheet[
                                                "$F" + cell[2:]].value
                                        except TypeError:
                                            practices = str(worksheet["$B" + cell[2:]].value)

        notes = worksheet["$X" + cell[2:]].value
        if notes is None:
            notes = "None"

        # logging.info("Practice: " + practices)
        if RorFcheck is None:
            value += 1
            continue
        elif RorFcheck is not None and pointsawarded != 0 and pointschosen is not None and pointschosen is not False:
            #print(pointsawarded)
            header.append((practices, pointsawarded, pointschosen, RorFcheck, errorcheck, practicetext, original_practicetext, notes))
            header_csv.append(('not mandatory',practices, pointsawarded, pointschosen, RorFcheck, errorcheck, practicetext,  notes))
            row += 1
        value += 1

    return header, header_csv


def dissect(file):
    try:
        workbook = openpyxl.load_workbook((file), data_only=True)
        worksheet = workbook['Verification Report']
    except:
        return
    basefilename = os.path.basename(file)
    if ".xlsm" in basefilename:
        year = "2012"
    elif ".xlsx" in basefilename:
        yearchecksheet = workbook['Info & Intro']
        yearcheck = yearchecksheet['$D$1'].value
        if "2015" in yearcheck:
            year = "2015"
        elif "2020" in yearcheck:
            year = "2020"
    # return year
    if worksheet["X5"].value is None:
        xlname = worksheet["X4"].value
        rOrF = worksheet["L5"].value
        goallevel = (worksheet["L1"].value,)
    else:
        xlname = worksheet["X4"].value + " " + str(worksheet["X5"].value)
        rOrF = worksheet["L5"].value
        goallevel = (worksheet["L1"].value,)
    try:
        xlname = xlname.replace(":", "")
        xlname = xlname.replace("/", "")
        xlname = xlname.replace("\\", "")
    except:
        return
    filename_extracted = rOrF + ' ' + xlname + ' extracted.xlsx'
    filename_reviewer = rOrF + ' ' + xlname + '_REVIEW.xlsx'
    value = 14
    row = 0
    pointstotal= 0



    header, header_csv = practiceextraction(value,row,worksheet)
    header2, header2_csv = mandatoryextraction(value,row,worksheet)
    #Excel workbook sheet for verifiers
    pworkbook = openpyxl.Workbook()
    psheet = pworkbook.active
    psheet.append(['Builder Opted Practices'])
    psheet.append(('Practice Number','Points','Description Added','R or F','Errors Present', 'Practice Text','Original Practice Text','Notes'))
    for topics, headerset in enumerate(header):
        # try:
        logging.info(headerset)
        psheet.append(headerset)
        pointstotal+= headerset[1]
        # except:
        #     continue
    psheet.append(('Total Points of Project:', pointstotal, 'Goal Level',goallevel[0], 'Min Points above ' ))
    psheet.append(['Mandatory Practices'])
    psheet.append(('Practice Number', '???', 'Practice Met?', 'R or F', 'Errors Present','Practice Text','Notes'))
    for topics, headerset in enumerate(header2):
        try:
            logging.info(headerset)
            psheet.append(headerset)
        except:
            continue



    fill1 = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')
    fill2 = PatternFill(start_color='FF9F9F', end_color='FF9F9F', fill_type='solid')
    fill3 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    psheet.conditional_formatting.add('$D$3:$D175',FormulaRule(formula=['=ISNUMBER(SEARCH("Rough/Final",D3))'], stopIfTrue=False,fill=fill1))
    psheet.conditional_formatting.add('$D$3:$D175',FormulaRule(formula=['=ISNUMBER(SEARCH("Final",D3))'], stopIfTrue=False,fill=fill2))
    psheet.conditional_formatting.add('$D$3:$D175',FormulaRule(formula=['=ISNUMBER(SEARCH("Rough",D3))'], stopIfTrue=False,fill=fill3))

    psheet.column_dimensions['a'].width = 20
    psheet.column_dimensions['B'].width = 10
    psheet.column_dimensions['C'].width = 18
    psheet.column_dimensions['D'].width = 15
    psheet.column_dimensions['E'].width = 25
    psheet.column_dimensions['F'].width = 35
    psheet.column_dimensions['G'].width = 35
    psheet.column_dimensions['H'].width = 35

    formatcell = len(header) + 4
    psheet.merge_cells('A1:G1')
    psheet.merge_cells('A'+str(formatcell)+':G'+str(formatcell))

    work_sheet_a1 = psheet['A1']
    work_sheet_a1.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=True)
    work_sheet_aNext = psheet['A'+str(formatcell)]
    work_sheet_aNext.font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=True)
    e_columns = ['A','B','C','D','E','F','G','H']
    for i in e_columns:
        work_sheet_a2 = psheet[i+'2']
        work_sheet_a2.font = Font(size=12, underline='single', color='CC9900', bold=True, italic=True)
    for i in e_columns:
        work_sheet_aman = psheet[i + str(formatcell+1)]
        work_sheet_aman.font = Font(size=12, underline='single', color='CC9900', bold=True, italic=True)
    pworkbook.save(os.path.dirname(os.path.realpath(file))+'/practice examiner/'+filename_extracted)

    #Excel workbook for verifiers end


    #CSV file for data extraction
    with open((os.path.dirname(os.path.realpath(file))+'/practice data analysis/'+filename_extracted)+ '.csv', 'w', newline='') as verifile:
        verifile_writer = csv.writer(verifile, delimiter=",")
        verifile_writer.writerow(('Practice Mandatory','Practice Number','Points','Attainment Note','Rough, Final, or Rough/Final',\
                                  'Point Availability','Practice Description', 'Notes','Certification Goal Level'))
        for topics, headerset in enumerate(header_csv):
             try:
                logging.info(headerset)
                headerset= headerset + goallevel
                verifile_writer.writerow(headerset)
             except:
                  continue

        for topics, headerset in enumerate(header2_csv):
            try:
                logging.info(headerset)
                headerset= headerset+ goallevel
                verifile_writer.writerow(headerset)
            except:
                continue
            #goal level included in title of data
    #CSV file for data extraction end
    print("Doine")
    return year, filename_reviewer






def macroYYYY(original_file, file):

    year, renamed_file = dissect(file)
    renamed_file = os.path.dirname(os.path.realpath(original_file))+'/'+renamed_file
    print(original_file)
    print(renamed_file)
    if year == "2012":
        print("lemon")
        return
    elif year == "2015":
        code = '''        
                sub marine()
    
                Dim i As Long, hasPhotoReviewWorksheet As Boolean
                Dim xlBook2  As Excel.Workbook
                Dim xlSheet2 As Excel.Worksheet
                Workbooks.Open filename:=''' + "\"" + original_file + "\"" + '''
                set xlbook = ActiveWorkbook
                For i = 1 To xlBook.Worksheets.Count
                    If xlBook.Worksheets(i).Name = "Photo Review" Then
                        Debug.Print xlBook.Worksheets(i).Name
                        hasPhotoReviewWorksheet = True
                        Exit For
                    End If
                Next i
    
                If hasPhotoReviewWorksheet Then
                    'Set xlSheet = xlBook.Worksheets("Photo Review")
                    'continue processing
                Else
                    xlBook.Unprotect ("4gotten!")
                    Set xlBook2 = xlApp.Workbooks.Open("y:\CustomSoftware\Databases\VerifiersCertifications\ExcelTemplates\2015NGBSNewConstructionPhotoReviewTemplate.xlsx")
                    xlBook2.Worksheets("Photo Review").select
                    xlBook2.Worksheets("Photo Review").Copy Before:=xlBook.Sheets("Rough Sig.")  '       xlBook.Sheets("Rough Sig.")
                    xlBook.Protect ("4gotten!")
                    xlBook2.Close False
                End If
    
                For Each ws In xlBook.Worksheets
                    mySheetName = ws.Name
    
                    Debug.Print mySheetName
    
                    Select Case mySheetName
                        Case "Photo Review"
    
                            If hasPhotoReviewWorksheet Then
                                xlBook.Unprotect ("4gotten!")
                                Set xlSheet = xlBook.Worksheets("Photo Review")
                                xlSheet.Visible = xlSheetVisible
                                xlSheet.Unprotect ("4gotten!")
    
                                xlBook.Worksheets("Photo Review").select
                                xlSheet.columns("H:I").EntireColumn.hidden = False
    
                            End If
    
                            xlSheet.Protect ("4gotten!")
                            xlBook.Protect ("4gotten!")
    
    
    
                        Case "Verification Report"
                        If mySheetName = "Verification Report" Then
                            Debug.Print mySheetName
                        End If
    
                        Set xlSheet = xlBook.Worksheets("Verification Report")
                        
                        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
                        ' The TryADifferentPassword on error was addedon 07/27/2018 because  a few workbooks have a bad password
                        '  on the the "Verifification Rpt WorkSheet" The password what applied in upper case when it shouldn't have.
                        '   on error try a different password and resume back at TryADifferentPassword:
                        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
                        xlSheet.Unprotect ("4gotten!")
                        xlSheet.Range("A:H").AutoFilter Field:=1
                        xlSheet.Range("A:H").AutoFilter Field:=7
                        xlSheet.Range("A:H").AutoFilter Field:=8
        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        TryADifferentPassword:
        ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    
                            xlSheet.columns("Z").EntireColumn.hidden = False
                            xlSheet.columns("X:Z").ColumnWidth = 22
    
                            'The reviewers wanted this unlocked so they could easily copy the text and send it
                            ' to a verifier if needed.
                            For r = 9 To 2500
                                If xlSheet.Range("L" & r).MergeCells Then
                                    'Debug.Print r
                                    For Each mrng In xlSheet.Range("L" & r)
                                        With mrng
                                            .MergeArea.Locked = False
                                        End With
                                    Next
                                Else
                                    xlSheet.Range("L" & r).Locked = False
                                End If
                            Next r
    
                            With xlSheet
                                .Protect Password:=("4gotten!"), AllowFiltering:=True
                                .EnableSelection = xlUnlockedCells
                            End With                      
                    End Select
                Next        
    
            ActiveWorkbook.SaveAs filename:=''' + "\"" + renamed_file + "\"" + ''', FileFormat:=xlOpenXMLWorkbook, CreateBackup:=False
            ThisWorkbook.Saved = True
    ActiveWorkbook.close
            end sub
        '''
    elif year == "2020":
        code = '''        
        sub marine()
        Dim i As Long, hasPhotoReviewWorksheet As Boolean
                Dim xlBook2  As Excel.Workbook
                Dim xlSheet2 As Excel.Worksheet
                Workbooks.Open filename:=''' + "\"" + original_file + "\"" + '''
                set xlbook = ActiveWorkbook
                 For i = 1 To xlBook.Worksheets.Count
                    If xlBook.Worksheets(i).Name = "Photo Review" Then
                        Debug.Print xlBook.Worksheets(i).Name
                        hasPhotoReviewWorksheet = True
                        Exit For
                    End If
                Next i
        For Each ws In xlBook.Worksheets
            mySheetName = ws.Name
            
            Debug.Print mySheetName
            
            Select Case mySheetName
                Case "Photo Review"
                    
                    If hasPhotoReviewWorksheet Then
                        xlBook.Unprotect ("4gotten!")
                        Set xlSheet = xlBook.Worksheets("Photo Review")
                        xlSheet.Visible = xlSheetVisible
                        xlSheet.Unprotect ("4gotten!")
                        
                        xlBook.Worksheets("Photo Review").select
                        xlSheet.columns("H:I").EntireColumn.hidden = False

                    End If
                    
                    'xlSheet.Protect ("4gotten!")
                    xlBook.Protect ("4gotten!")
                    
                    
                
                Case "Verification Report"
                
                    If mySheetName = "Verification Report" Then
                        Debug.Print mySheetName
                    End If
                    
                    ' *************************************************************
                    ' Set the worksheet variable to "Verification Report" worksheet
                    ' *************************************************************
                    Set xlSheet = xlBook.Worksheets("Verification Report")
                    
                    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
                    ' The TryADifferentPassword on error was addedon 07/27/2018 because  a few workbooks have a bad password
                    '  on the the "Verifification Rpt WorkSheet" The password what applied in upper case when it shouldn't have.
                    '   on error try a different password and resume back at TryADifferentPassword:
                    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
                    xlSheet.Unprotect ("4gotten!")
                    xlSheet.Range("A:H").AutoFilter Field:=1
                    xlSheet.Range("A:H").AutoFilter Field:=7
                    xlSheet.Range("A:H").AutoFilter Field:=8
                
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
TryADifferentPassword:
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

                    xlSheet.columns("Z:AA").EntireColumn.hidden = False
                    xlSheet.columns("X:Z").ColumnWidth = 22
                    xlSheet.columns("AA").ColumnWidth = 6
                    
                    
                    
                    'The reviewers wanted this unlocked so they could easily copy the text and send it
                    '   to a verifier if needed.
                    For r = 9 To 2500
                        If xlSheet.Range("L" & r).MergeCells Then
                            'Debug.Print r
                            For Each mrng In xlSheet.Range("L" & r)
                                With mrng
                                    .MergeArea.Locked = False
                                End With
                            Next
                        Else
                            xlSheet.Range("L" & r).Locked = False
                        End If
                    Next r
                    

                    
                    'The reviewers wanted this unlocked so they could easily copy the text and send it
                    '   to a verifier if needed.
                    For r = 13 To 2500
                        If xlSheet.Range("Z" & r).MergeCells Then
                            For Each mrng In xlSheet.Range("Z" & r)
                                With mrng
                                    .MergeArea.Locked = False
                                End With
                            Next
                        Else
                            xlSheet.Range("Z" & r).Locked = False
                        End If
                    Next r

                    
                    
                    
                    With xlSheet
                        .Protect Password:=("4gotten!"), AllowFiltering:=True
                        .EnableSelection = xlUnlockedCells
                    End With
                    
            End Select
        Next
        

    ActiveWorkbook.SaveAs filename:=''' + "\"" + renamed_file + "\"" + ''', FileFormat:=xlOpenXMLWorkbook, CreateBackup:=False
    ThisWorkbook.Saved = True
    ActiveWorkbook.close
            end sub
'''
    return code


def task():
    #xl.Quit()
    directory_2015 = os.path.dirname(os.path.dirname(os.path.dirname(os.getcwd())))
    logging.basicConfig(stream=sys.stderr, level=logging.INFO)
    print(directory_2015)
    for files in glob.glob(directory_2015 + "/*.xlsx", recursive=False):
        print(files)
        print(files[:-5])
        reviewfiles = glob.glob(files[:-5] + "*_REVIEW.xlsx", recursive=False)
        # check if any other files have the same code extension as other files
        creation = os.path.getmtime(files)
        print(creation)
        print(reviewfiles)
        print("evaluating")
        if "_REVIEW" not in files and not reviewfiles:
            # #if there is no reviewer file with the same code name, then create and name the
            # reviewer version of the file
            print("now running")
            original_file = files
            renamed_file = files[:-5] + "_REVIEW.xlsx"
            ss = xl.Workbooks.Open(os.path.abspath("excelsheet3.xlsm"), ReadOnly=1)
            try:
                # remove the module from the 2nd derivative file
                for i in ss.VBProject.VBComponents:
                    old_xlmodule = ss.VBProject.VBComponents(i.Name)
                    if old_xlmodule.Type in [1, 2, 3]:
                        ss.VBProject.VBComponents.Remove(old_xlmodule)
            except: 
                print("module does not exist")
            # add the module
            print("adding module")
            xlmodule = ss.VBProject.VBComponents.Add(1)

            code = macroYYYY(original_file,  files)

            #use the appropriate code as the macro base and run the macro
            xlmodule.CodeModule.AddFromString(code)
            xlmodule.Name = 'translate'
            ss.Application.Run('translate.marine')
            ss.Close(SaveChanges=False)
            #xl.Quit()
            os.remove(files)
            # delete the fileafter exiting excel
        else:
            print(reviewfiles)
    root.after(5000, task)

def endbutton():
    print("end")

xl = win32.gencache.EnsureDispatch('Excel.Application')
xl.Quit()
xl = win32.gencache.EnsureDispatch('Excel.Application')
xl.Visible = False

root = tk.Tk()

root.after(5000, task)
root.mainloop()

